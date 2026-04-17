"""
Import all Excel sheets into PostgreSQL tables.
Parses: P2P Links, B2B Links, FW Links, Server AS, IP Range,
        MLAG, MSTP, VLANs from switch_guide.xlsx.

Usage:
    python parser/import_all_sheets.py switch_guide.xlsx
    # Or with explicit DSN:
    python parser/import_all_sheets.py switch_guide.xlsx --dsn "postgresql://..."
"""

from __future__ import annotations
import argparse
import os
import sys

try:
    import psycopg2
except ImportError:
    print("pip install psycopg2-binary")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


DEFAULT_DSN = "postgresql://switchbuilder:switchbuilder@localhost:5432/switchbuilder"


def _str(val) -> str:
    """Convert cell value to stripped string, or empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


def _headers(ws) -> list[str]:
    """Read first row as lowercase header strings."""
    return [_str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]


def _rows(ws, headers: list[str]) -> list[dict]:
    """Yield each data row as a dict keyed by header name."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, h in enumerate(headers):
            if i < len(row) and h:
                d[h] = _str(row[i]) if i < len(row) else ""
            elif h:
                d[h] = ""
        # skip completely empty rows
        if any(v for v in d.values()):
            rows.append(d)
    return rows


def import_p2p(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM p2p_links")
    for r in rows:
        cur.execute("""
            INSERT INTO p2p_links (region, building, link_id, vlan, device_a, port_a, device_a_ip, device_b, port_b, device_b_ip, subnet, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (r.get('region',''), r.get('building',''), r.get('link id',''), r.get('vlan',''),
              r.get('device a',''), r.get('port a',''), r.get('device a ip',''),
              r.get('device b',''), r.get('port b',''), r.get('device b ip',''),
              r.get('subnet',''), r.get('status','')))
    print(f"  P2P Links: {len(rows)} rows")


def import_b2b(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM b2b_links")
    for r in rows:
        cur.execute("""
            INSERT INTO b2b_links (link_id, vlan, building_a, device_a, port_a, module_a, device_a_ip,
                building_b, device_b, port_b, module_b, device_b_ip, tx, rx, media, speed, subnet, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (r.get('link id',''), r.get('vlan',''), r.get('building a',''),
              r.get('device a',''), r.get('port a',''), r.get('module a',''), r.get('device a ip',''),
              r.get('building b',''), r.get('device b',''), r.get('port b',''),
              r.get('module b',''), r.get('device b ip',''),
              r.get('tx',''), r.get('rx',''), r.get('media',''), r.get('speed',''),
              r.get('subnet',''), r.get('status','')))
    print(f"  B2B Links: {len(rows)} rows")


def import_fw(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM fw_links")
    for r in rows:
        cur.execute("""
            INSERT INTO fw_links (building, link_id, vlan, switch, switch_port, switch_ip,
                firewall, firewall_port, firewall_ip, subnet, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (r.get('building',''), r.get('link id',''), r.get('vlan',''),
              r.get('switch',''), r.get('switch port',''), r.get('switch ip',''),
              r.get('firewall',''), r.get('firewall port',''), r.get('firewall ip',''),
              r.get('subnet',''), r.get('status','')))
    print(f"  FW Links: {len(rows)} rows")


def import_server_as(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM server_as")
    for r in rows:
        cur.execute("""
            INSERT INTO server_as (building, server_as, status)
            VALUES (%s,%s,%s)
        """, (r.get('building',''), r.get('server as',''), r.get('status','')))
    print(f"  Server AS: {len(rows)} rows")


def import_ip_range(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM ip_ranges")
    for r in rows:
        cur.execute("""
            INSERT INTO ip_ranges (region, pool_name, block, purpose, notes, status)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (r.get('region',''), r.get('ip pool name',''), r.get('block',''),
              r.get('purpose',''), r.get('notes',''), r.get('status','')))
    print(f"  IP Ranges: {len(rows)} rows")


def import_mlag(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM mlag_config")
    for r in rows:
        cur.execute("""
            INSERT INTO mlag_config (building, domain_type, mlag_domain, switch_a, switch_b,
                b2b_partner, status, peer_link_ae, physical_members, peer_vlan, trunk_vlans,
                shared_domain_mac, peer_link_subnet, node0_ip, node1_ip, node0_ip_link2, node1_ip_link2, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (r.get('building',''), r.get('domain type',''), r.get('mlag domain',''),
              r.get('switch a (ae 1-99)',''), r.get('switch b (ae 100-199)',''),
              r.get('b2b partner',''), r.get('status',''),
              r.get('peer-link ae',''), r.get('physical members',''),
              r.get('peer vlan',''), r.get('trunk vlans',''),
              r.get('shared domain mac',''), r.get('peer-link subnet',''),
              r.get('node 0 ip',''), r.get('node 1 ip',''),
              r.get('node 0 ip (link 2)',''), r.get('node 1 ip (link 2)',''),
              r.get('notes','')))
    print(f"  MLAG Config: {len(rows)} rows")


def import_mstp(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM mstp_config")
    for r in rows:
        cur.execute("""
            INSERT INTO mstp_config (building, device_name, device_role, mstp_priority, notes, status)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (r.get('building',''), r.get('device name',''), r.get('device role',''),
              r.get('mstp priority',''), r.get('notes',''), r.get('status','')))
    print(f"  MSTP Config: {len(rows)} rows")


def import_vlans(cur, ws):
    headers = _headers(ws)
    rows = _rows(ws, headers)
    cur.execute("DELETE FROM vlan_inventory")
    for r in rows:
        cur.execute("""
            INSERT INTO vlan_inventory (block, vlan_id, name, network_address, subnet, gateway, usable_range, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (r.get('/21 block',''), r.get('vlan',''), r.get('name',''),
              r.get('network address',''), r.get('subnet',''), r.get('gateway',''),
              r.get('usable host range',''), r.get('status','')))
    print(f"  VLANs: {len(rows)} rows")


# Sheet name -> import function mapping
def import_servers(cur, wb):
    """Import all per-building server sheets (MEP-91 Servers, GBG Servers, etc.)."""
    cur.execute("DELETE FROM servers")
    total = 0
    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith(" Servers"):
            continue
        ws = wb[sheet_name]
        rows_list = list(ws.iter_rows(values_only=True))
        if len(rows_list) < 10:
            continue

        # Row 1: Building name in col B
        building = _str(rows_list[0][1]) if len(rows_list[0]) > 1 else ""
        if not building:
            building = sheet_name.replace(" Servers", "")

        # Data rows start at row 10 (index 9), headers at row 9 (index 8)
        # Column layout (20 cols): ServerName, ServerAS, Loopback,
        #   NIC1_IP, NIC1_Router, NIC1_Subnet, NIC1_Status,
        #   NIC2_IP, NIC2_Router, NIC2_Subnet, NIC2_Status,
        #   NIC3_IP, NIC3_Router, NIC3_Subnet, NIC3_Status,
        #   NIC4_IP, NIC4_Router, NIC4_Subnet, NIC4_Status, Status
        for row in rows_list[9:]:  # skip header rows
            vals = list(row) + [None] * 20  # pad to avoid index errors
            server_name = _str(vals[0])
            if not server_name:
                continue
            # Last non-None value is the overall status
            status = _str(vals[19]) if vals[19] else _str(vals[15]) if vals[15] else "RESERVED"
            # For rows with NIC status columns (first 10 servers have them at pos 6,10,14,18)
            cur.execute("""
                INSERT INTO servers (building, server_name, server_as, loopback_ip,
                    nic1_ip, nic1_router, nic1_subnet, nic1_status,
                    nic2_ip, nic2_router, nic2_subnet, nic2_status,
                    nic3_ip, nic3_router, nic3_subnet, nic3_status,
                    nic4_ip, nic4_router, nic4_subnet, nic4_status, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (building, server_name, _str(vals[1]), _str(vals[2]),
                  _str(vals[3]), _str(vals[4]), _str(vals[5]), _str(vals[6]),
                  _str(vals[7]), _str(vals[8]), _str(vals[9]), _str(vals[10]),
                  _str(vals[11]), _str(vals[12]), _str(vals[13]), _str(vals[14]),
                  _str(vals[15]), _str(vals[16]), _str(vals[17]), _str(vals[18]),
                  status))
            total += 1
        print(f"  {sheet_name}: {sum(1 for r in rows_list[9:] if r[0])} servers")
    print(f"  Servers total: {total} rows")


def import_asn_loop(cur, ws):
    """Import ASN & Loopback sheet — updates switch_guide rows and populates asn_definitions."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    updated = 0
    for r in rows:
        if not r or not r[2]:  # col C = Device Name
            continue
        device_name = _str(r[2])
        asn         = _str(r[4])   # col E
        loopback_ip = _str(r[5])   # col F
        loopback_sub= _str(r[6])   # col G
        mgmt_ip     = _str(r[11])  # col L
        mgmt_l3     = _str(r[12])  # col M
        primary_ip  = _str(r[10])  # col K
        status      = _str(r[9])   # col J
        device_type = _str(r[3])   # col D
        region      = _str(r[0])   # col A
        building    = _str(r[1])   # col B
        # Update switch_guide with ASN + loopback + management + status data
        cur.execute("""
            UPDATE switch_guide SET
                asn = %s,
                loopback_ip = COALESCE(%s::inet, loopback_ip),
                loopback_subnet = COALESCE(%s::cidr, loopback_subnet),
                management_ip = COALESCE(%s::inet, management_ip),
                mgmt_l3_ip = COALESCE(%s::inet, mgmt_l3_ip),
                primary_ip = %s,
                status = COALESCE(NULLIF(%s, ''), status),
                device_type = COALESCE(NULLIF(%s, ''), device_type),
                region = COALESCE(NULLIF(%s, ''), region)
            WHERE switch_name = %s
        """, (asn,
              loopback_ip or None,
              loopback_sub or None,
              mgmt_ip or None,
              mgmt_l3 or None,
              primary_ip,
              status, device_type, region,
              device_name))
        updated += cur.rowcount
        # Upsert ASN definition
        if asn:
            asn_type = 'Building Router'
            if 'firewall' in device_type.lower(): asn_type = 'Firewall'
            elif 'storage' in device_type.lower(): asn_type = 'Server'
            elif 'reserved' in device_type.lower(): asn_type = 'Reserved'
            desc = f"{building} {device_type} — {device_name}"
            cur.execute("""
                INSERT INTO asn_definitions (asn, description, asn_type)
                VALUES (%s, %s, %s)
                ON CONFLICT (asn) DO UPDATE SET description = EXCLUDED.description, asn_type = EXCLUDED.asn_type
            """, (asn, desc, asn_type))
    print(f"  ASN & Loop: {updated} switch_guide rows updated")


SHEET_MAP = {
    "asn & loop":  import_asn_loop,
    "p2p links":   import_p2p,
    "b2b links":   import_b2b,
    "fw links":    import_fw,
    "server as":   import_server_as,
    "ip range":    import_ip_range,
    "mlag":        import_mlag,
    "mstp":        import_mstp,
    "vlan's":      import_vlans,
    "vlans":       import_vlans,
}


def main():
    parser = argparse.ArgumentParser(description="Import all Excel sheets into DB")
    parser.add_argument("xlsx", help="Path to switch_guide.xlsx")
    parser.add_argument("--dsn", default=os.environ.get("SWITCHBUILDER_DSN", DEFAULT_DSN))
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    print(f"Opened: {args.xlsx}")
    print(f"Sheets: {wb.sheetnames}")

    conn = psycopg2.connect(args.dsn)
    cur = conn.cursor()

    imported = 0
    for sheet_name in wb.sheetnames:
        key = sheet_name.strip().lower()
        if key in SHEET_MAP:
            print(f"\nImporting '{sheet_name}'...")
            SHEET_MAP[key](cur, wb[sheet_name])
            imported += 1

    # Import per-building server sheets
    print("\nImporting Server sheets...")
    import_servers(cur, wb)
    imported += 1

    conn.commit()
    cur.close()
    conn.close()
    print(f"\nDone — imported {imported} sheets")


if __name__ == "__main__":
    main()
