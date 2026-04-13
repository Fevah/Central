"""
Switch Guide Excel Importer
============================
Imports switch_guide.xlsx into structured Python dicts and
optionally inserts them into PostgreSQL.

Expected Excel structure (flexible - adapts to column names found):
  Typical columns in switch_guide.xlsx:
    - Switch Name / Hostname
    - Site / Building / Location
    - Floor / Rack
    - Model / Part Number
    - Serial Number
    - Management IP
    - Uplink Switch / Uplink Port
    - Port / Connection details (one row per connection)
    - Enabled / Active (boolean)
    - Notes

The importer auto-detects column names using fuzzy matching so it
handles slight naming variations between spreadsheet versions.

Usage:
    importer = ExcelImporter("switch_guide.xlsx")
    switches, connections = importer.load()
    for sw in switches:
        print(sw)

Dependencies:
    pip install openpyxl
"""

from __future__ import annotations

import re
import os
from dataclasses import dataclass, field
from typing import Optional, Any


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class GuideSwitch:
    """One switch record from the guide spreadsheet."""
    switch_name:    Optional[str] = None
    site:           Optional[str] = None
    building:       Optional[str] = None
    floor:          Optional[str] = None
    rack:           Optional[str] = None
    model:          Optional[str] = None
    serial_number:  Optional[str] = None
    management_ip:  Optional[str] = None
    uplink_switch:  Optional[str] = None
    uplink_port:    Optional[str] = None
    notes:          Optional[str] = None
    enabled:        bool = True
    raw_data:       dict = field(default_factory=dict)  # all original columns


@dataclass
class GuideConnection:
    """One port/connection record from the guide spreadsheet."""
    switch_name:      Optional[str] = None
    local_port:       Optional[str] = None
    remote_device:    Optional[str] = None
    remote_port:      Optional[str] = None
    connection_type:  Optional[str] = None   # trunk, access, uplink, downlink
    vlan_id:          Optional[int] = None   # native/data VLAN
    voice_vlan_id:    Optional[int] = None   # voice VLAN
    vlan_members:     Optional[str] = None   # trunk allowed VLANs, e.g. "1-310"
    speed:            Optional[str] = None   # port speed in Mbps
    description:      Optional[str] = None   # interface description override
    enabled:          bool = True
    notes:            Optional[str] = None
    raw_data:         dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Column name normaliser
# ---------------------------------------------------------------------------

# Map of canonical field name -> list of possible column header variants
COLUMN_ALIASES: dict[str, list[str]] = {
    "switch_name":   ["switch name", "hostname", "switch", "device name", "name", "device"],
    "site":          ["site", "location", "region", "facility"],
    "building":      ["building", "bldg", "building name"],
    "floor":         ["floor", "level"],
    "rack":          ["rack", "rack id", "rack name", "rack unit"],
    "model":         ["model", "model number", "part number", "pn", "hardware model"],
    "serial_number": ["serial", "serial number", "s/n", "sn"],
    "management_ip": ["management ip", "mgmt ip", "ip address", "ip", "mgmt", "management"],
    "uplink_switch": ["uplink switch", "uplink device", "connected to", "upstream switch", "parent switch"],
    "uplink_port":   ["uplink port", "upstream port", "connected port", "trunk port"],
    "notes":         ["notes", "comments", "description", "remark"],
    "enabled":       ["enabled", "active", "status", "in service"],
    # Connection-specific
    "local_port":    ["local port", "port", "interface", "local interface", "switch port", "port number"],
    "remote_device": ["remote device", "connected device", "far end device", "peer device", "device connected", "device"],
    "remote_port":   ["remote port", "far end port", "peer port", "connected port"],
    "connection_type": ["connection type", "type", "port type", "link type", "mode"],
    "vlan_id":       ["vlan", "vlan id", "data vlan", "native vlan", "access vlan", "pvid", "vlanid"],
    "voice_vlan_id": ["voice vlan", "voice vlan id", "voip vlan", "voice", "vvid"],
    "vlan_members":  ["vlan members", "allowed vlans", "trunk vlans", "tagged vlans", "vlans"],
    "speed":         ["speed", "port speed", "link speed", "bandwidth"],
    "description":   ["description", "port description", "label", "port label"],
}

def _normalise_header(header: str) -> str:
    """Lower-case and strip a column header for comparison."""
    return re.sub(r'\s+', ' ', str(header).strip().lower())

def _map_columns(headers: list[str]) -> dict[str, int]:
    """
    Returns {canonical_field: column_index} for all recognised headers.
    """
    normalised = [_normalise_header(h) for h in headers]
    result: dict[str, int] = {}

    for field_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            try:
                idx = normalised.index(alias)
                result[field_name] = idx
                break
            except ValueError:
                continue
    return result


# ---------------------------------------------------------------------------
# Excel Importer
# ---------------------------------------------------------------------------

class ExcelImporter:
    """
    Reads switch_guide.xlsx and produces lists of GuideSwitch and
    GuideConnection objects.

    Sheet detection:
      - If there is a sheet named 'switches' or 'inventory', that is used
        for GuideSwitch records.
      - If there is a sheet named 'connections' or 'ports', that is used
        for GuideConnection records.
      - If only one sheet exists, it is parsed as switch inventory and
        also scanned for connection columns.
    """

    def __init__(self, filepath: str):
        self.filepath = filepath
        self._switches: list[GuideSwitch] = []
        self._connections: list[GuideConnection] = []

    # -----------------------------------------------------------------------
    # Public
    # -----------------------------------------------------------------------

    def load(self) -> tuple[list[GuideSwitch], list[GuideConnection]]:
        """
        Parse the Excel file and return (switches, connections).
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

        switch_sheet = self._find_sheet(
            wb, sheet_names_lower,
            ["switches", "inventory", "switch inventory", "devices"]
        )
        connection_sheet = self._find_sheet(
            wb, sheet_names_lower,
            ["connections", "ports", "port connections", "cabling", "links"]
        )

        if switch_sheet is None and connection_sheet is None:
            # Fall back to first sheet
            switch_sheet = wb[wb.sheetnames[0]]

        if switch_sheet is not None:
            self._parse_switch_sheet(switch_sheet)

        if connection_sheet is not None and connection_sheet != switch_sheet:
            self._parse_connection_sheet(connection_sheet)

        wb.close()
        return self._switches, self._connections

    # -----------------------------------------------------------------------
    # Sheet finders
    # -----------------------------------------------------------------------

    def _find_sheet(self, wb, sheet_names_lower: dict, candidates: list[str]):
        for name in candidates:
            if name in sheet_names_lower:
                return wb[sheet_names_lower[name]]
        return None

    # -----------------------------------------------------------------------
    # Switch sheet parser
    # -----------------------------------------------------------------------

    def _parse_switch_sheet(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        # Find header row (first row with non-empty cells)
        header_row_idx = 0
        for i, row in enumerate(rows):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if non_empty:
                header_row_idx = i
                break

        headers = [str(c) if c is not None else "" for c in rows[header_row_idx]]
        col_map = _map_columns(headers)

        has_connection_cols = (
            "local_port" in col_map or "remote_device" in col_map
        )

        for row in rows[header_row_idx + 1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue  # skip blank rows

            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            # Build GuideSwitch
            sw = GuideSwitch(raw_data=raw)
            sw.switch_name   = self._col_val(row, col_map, "switch_name")
            sw.site          = self._col_val(row, col_map, "site")
            sw.building      = self._col_val(row, col_map, "building")
            sw.floor         = self._col_val(row, col_map, "floor")
            sw.rack          = self._col_val(row, col_map, "rack")
            sw.model         = self._col_val(row, col_map, "model")
            sw.serial_number = self._col_val(row, col_map, "serial_number")
            sw.management_ip = self._col_val(row, col_map, "management_ip")
            sw.uplink_switch = self._col_val(row, col_map, "uplink_switch")
            sw.uplink_port   = self._col_val(row, col_map, "uplink_port")
            sw.notes         = self._col_val(row, col_map, "notes")
            sw.enabled       = self._parse_bool(
                self._col_val(row, col_map, "enabled"), default=True
            )

            # Skip rows with no identifying info
            if sw.switch_name is None and sw.management_ip is None:
                continue

            self._switches.append(sw)

            # If connection columns present in same sheet, extract them too
            if has_connection_cols:
                conn = GuideConnection(raw_data=raw)
                conn.switch_name     = sw.switch_name
                conn.local_port      = self._col_val(row, col_map, "local_port")
                conn.remote_device   = self._col_val(row, col_map, "remote_device")
                conn.remote_port     = self._col_val(row, col_map, "remote_port")
                conn.connection_type = self._col_val(row, col_map, "connection_type")
                conn.vlan_id         = self._parse_int(self._col_val(row, col_map, "vlan_id"))
                conn.voice_vlan_id   = self._parse_int(self._col_val(row, col_map, "voice_vlan_id"))
                conn.vlan_members    = self._col_val(row, col_map, "vlan_members")
                conn.speed           = self._col_val(row, col_map, "speed")
                conn.description     = self._col_val(row, col_map, "description")
                conn.enabled         = sw.enabled
                conn.notes           = sw.notes
                if conn.local_port or conn.remote_device:
                    self._connections.append(conn)

    # -----------------------------------------------------------------------
    # Connection sheet parser
    # -----------------------------------------------------------------------

    def _parse_connection_sheet(self, sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        header_row_idx = 0
        for i, row in enumerate(rows):
            if any(c is not None and str(c).strip() for c in row):
                header_row_idx = i
                break

        headers = [str(c) if c is not None else "" for c in rows[header_row_idx]]
        col_map = _map_columns(headers)

        for row in rows[header_row_idx + 1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue

            raw = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            conn = GuideConnection(raw_data=raw)
            conn.switch_name     = self._col_val(row, col_map, "switch_name")
            conn.local_port      = self._col_val(row, col_map, "local_port")
            conn.remote_device   = self._col_val(row, col_map, "remote_device")
            conn.remote_port     = self._col_val(row, col_map, "remote_port")
            conn.connection_type = self._col_val(row, col_map, "connection_type")
            conn.vlan_id         = self._parse_int(self._col_val(row, col_map, "vlan_id"))
            conn.voice_vlan_id   = self._parse_int(self._col_val(row, col_map, "voice_vlan_id"))
            conn.vlan_members    = self._col_val(row, col_map, "vlan_members")
            conn.speed           = self._col_val(row, col_map, "speed")
            conn.description     = self._col_val(row, col_map, "description")
            conn.enabled         = self._parse_bool(
                self._col_val(row, col_map, "enabled"), default=True
            )
            conn.notes           = self._col_val(row, col_map, "notes")

            if conn.switch_name or conn.local_port:
                self._connections.append(conn)

    # -----------------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------------

    def _col_val(self, row: tuple, col_map: dict, field_name: str) -> Optional[str]:
        idx = col_map.get(field_name)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def _parse_int(self, val: Any) -> Optional[int]:
        if val is None:
            return None
        try:
            return int(str(val).strip())
        except (ValueError, TypeError):
            return None

    def _parse_bool(self, val: Any, default: bool = True) -> bool:
        if val is None:
            return default
        s = str(val).strip().lower()
        if s in ("true", "yes", "1", "y", "enabled", "active", "x"):
            return True
        if s in ("false", "no", "0", "n", "disabled", "inactive", ""):
            return False
        return default

    # -----------------------------------------------------------------------
    # Sheet inspection helper
    # -----------------------------------------------------------------------

    def inspect(self) -> dict:
        """
        Returns metadata about the Excel file without fully parsing it.
        Useful for understanding an unknown workbook structure.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required: pip install openpyxl")

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        info = {"sheets": {}}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=3))
            info["sheets"][sheet_name] = {
                "max_row":    ws.max_row,
                "max_column": ws.max_column,
                "headers":    [str(c) for c in rows[0]] if rows else [],
            }
        wb.close()
        return info


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    import json

    filepath = sys.argv[1] if len(sys.argv) > 1 else "switch_guide.xlsx"

    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)

    importer = ExcelImporter(filepath)

    # First, inspect the workbook structure
    print("=== Workbook Inspection ===")
    info = importer.inspect()
    print(json.dumps(info, indent=2, default=str))
    print()

    # Then load
    switches, connections = importer.load()
    print(f"=== Loaded {len(switches)} switches, {len(connections)} connections ===")

    for sw in switches[:5]:
        print(f"  Switch: {sw.switch_name}  site={sw.site}  ip={sw.management_ip}  model={sw.model}")

    if connections:
        print(f"\nFirst 5 connections:")
        for conn in connections[:5]:
            print(f"  {conn.switch_name} port={conn.local_port} -> {conn.remote_device} {conn.remote_port}")
